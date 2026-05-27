import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Sum

import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from apps.institutions.models import Institution, BudgetLimit
from apps.budget_requests.models import BudgetRequest, RequestStatus, RequestLine
from apps.monitoring.models import FactPayment
from apps.classifiers.models import KOSGU


# ─── ПФХД canonical expense line items ──────────────────────────────────────
# (display_num, kvr_code, kosgu_code, name, row_code)
PFHD_LINES = [
    (1,  '111', '211', 'Заработная плата работников учреждения',          '101'),
    (2,  '119', '213', 'Взносы по обязательному социальному страхованию', '102'),
    (3,  '244', '221', 'Услуги связи',                                    '103'),
    (4,  '244', '222', 'Транспортные услуги',                             '104'),
    (5,  '244', '223', 'Коммунальные услуги',                             '105'),
    (6,  '244', '224', 'Арендная плата за пользование имуществом',        '106'),
    (7,  '244', '225', 'Работы, услуги по содержанию имущества',          '107'),
    (8,  '244', '226', 'Прочие работы и услуги',                          '108'),
    (9,  '244', '290', 'Прочие расходы (налоги, сборы, взносы)',          '109'),
    (10, '244', '310', 'Увеличение стоимости основных средств',           '110'),
    (11, '244', '340', 'Увеличение стоимости материальных запасов',       '111'),
    (12, '244', '296', 'Иные расходы (культурные мероприятия)',           '112'),
    (13, '112', '211', 'Заработная плата: иные выплаты персоналу',        '113'),
]

KVR_GROUPS = [
    ('111', 'КВР 111 — Фонд оплаты труда и страховые взносы'),
    ('119', 'КВР 119 — Взносы по обязательному страхованию'),
    ('244', 'КВР 244 — Прочие закупки товаров, работ и услуг'),
    ('112', 'КВР 112 — Иные выплаты персоналу'),
]


def _get_institution_pfhd(institution: Institution, year: int) -> dict:
    """Collect ПФХД data for a single institution for the given year."""
    limit_obj = BudgetLimit.objects.filter(institution=institution, year=year).first()
    total_limit = limit_obj.total_limit if limit_obj else Decimal('0')

    # Revenue totals by funding type (from approved/included requests)
    approved_qs = BudgetRequest.objects.filter(
        institution=institution,
        period_year=year,
        status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED],
    )

    def _rev(ft: str) -> Decimal:
        s = approved_qs.filter(funding_type=ft).aggregate(
            t=Sum('approved_amount')
        )['t']
        if not s:
            s = approved_qs.filter(funding_type=ft).aggregate(
                t=Sum('total_amount')
            )['t']
        return s or Decimal('0')

    rev_subsidy = _rev('subsidy')
    rev_other   = _rev('other')
    rev_paid    = _rev('paid')
    total_rev   = rev_subsidy + rev_other + rev_paid

    # If no approved requests, fall back to the limit
    if total_rev == 0 and total_limit > 0:
        rev_subsidy = total_limit
        total_rev   = total_limit

    # Expense amounts grouped by (kvr_code, kosgu_code)
    lines_qs = RequestLine.objects.filter(
        request__institution=institution,
        request__period_year=year,
        request__status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED],
    ).select_related('kosgu', 'kvr', 'request')

    amounts: dict = {}
    for line in lines_qs:
        kvr_code   = line.kvr.code if line.kvr else '244'
        kosgu_code = line.kosgu.code
        key = (kvr_code, kosgu_code)
        if key not in amounts:
            amounts[key] = {k: Decimal('0') for k in ('total', 'q1', 'q2', 'q3', 'q4')}
        amounts[key]['total'] += line.amount

        req = line.request
        if req.period_type == 'quarterly' and req.period_quarter:
            amounts[key][f'q{req.period_quarter}'] += line.amount
        else:
            per_q = line.amount / 4
            for q in ('q1', 'q2', 'q3', 'q4'):
                amounts[key][q] += per_q

    # Fact payments grouped by KOSGU
    fact_qs = FactPayment.objects.filter(institution=institution, year=year)
    fact_by_kosgu: dict = {}
    total_fact = Decimal('0')
    for fp in fact_qs:
        code = fp.kosgu.code
        fact_by_kosgu[code] = fact_by_kosgu.get(code, Decimal('0')) + fp.amount
        total_fact += fp.amount

    return {
        'institution': institution,
        'total_limit': total_limit,
        'rev_subsidy': rev_subsidy,
        'rev_other':   rev_other,
        'rev_paid':    rev_paid,
        'total_rev':   total_rev,
        'amounts':     amounts,
        'fact_by_kosgu': fact_by_kosgu,
        'total_fact':  total_fact,
    }


# ─── Excel style helpers ─────────────────────────────────────────────────────
def _thin_border() -> Border:
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)


def _header_border() -> Border:
    t = Side(style='medium', color='888888')
    return Border(left=t, right=t, top=t, bottom=t)


FILL_TITLE   = PatternFill('solid', fgColor='1F3864')   # dark navy
FILL_SECTION = PatternFill('solid', fgColor='D6E4F7')   # light blue section
FILL_KVR_GRP = PatternFill('solid', fgColor='EEF2F8')   # pale blue KVR group
FILL_SUBTOT  = PatternFill('solid', fgColor='F2F4F8')   # very light gray subtotal
FILL_TOTAL   = PatternFill('solid', fgColor='D9E1EE')   # medium blue-gray total
FILL_COLHEAD = PatternFill('solid', fgColor='2E5E9E')   # column header navy

FONT_WHITE_BOLD = Font(bold=True, color='FFFFFF', size=10, name='Calibri')
FONT_BOLD       = Font(bold=True, size=10, name='Calibri')
FONT_NORMAL     = Font(size=10, name='Calibri')
FONT_SMALL_GREY = Font(size=9, color='666666', name='Calibri')

ALIGN_CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_CENTER_TOP = Alignment(horizontal='center', vertical='top', wrap_text=True)
ALIGN_LEFT       = Alignment(horizontal='left',  vertical='center', wrap_text=True)
ALIGN_RIGHT      = Alignment(horizontal='right', vertical='center')

NUM_RUB  = '#,##0.00'
NUM_PCT  = '0.0%'
NUM_INT  = '0'

COL_WIDTHS = [5, 7, 7, 40, 8, 14, 11, 11, 11, 11, 13, 7, 14, 7]


def _set_col_widths(ws) -> None:
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_title_sheet(ws, year: int, today: date) -> None:
    """Write the Титул (title page) sheet."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 60

    def row(text: str, bold: bool = False, size: int = 11, color: str = '111111') -> None:
        r = ws.max_row + 1 if ws.max_row else 1
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color, name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 10
    data = [
        ('УТВЕРЖДАЮ', True, 12, '111111'),
        ('Руководитель Управления культуры', False, 11, '111111'),
        ('администрации г. Сочи', False, 11, '111111'),
        ('________________ /________________/', False, 11, '555555'),
        (f'«___» ____________ {year} г.', False, 11, '555555'),
        ('', False, 8, ''),
        ('ПЛАН ФИНАНСОВО-ХОЗЯЙСТВЕННОЙ ДЕЯТЕЛЬНОСТИ', True, 16, '1F3864'),
        (f'на {year} год и плановый период {year + 1} года', True, 13, '1F3864'),
        ('', False, 8, ''),
        ('Муниципальное казённое учреждение', False, 11, '111111'),
        ('«Централизованная бухгалтерия Управления культуры', False, 11, '111111'),
        ('администрации муниципального образования', False, 11, '111111'),
        ('городского округа город-курорт Сочи Краснодарского края»', False, 11, '111111'),
        ('', False, 8, ''),
        ('ИНН 2320035700  /  КПП 232001001', False, 10, '555555'),
        ('ОКТМО 03726000001   ОКПО 51234567   ОКОГУ 1300214', False, 10, '555555'),
        ('', False, 8, ''),
        ('Единица измерения: рублей (коп.)', False, 10, '333333'),
        ('', False, 8, ''),
        ('Сформировано информационной системой управления бюджетированием МКУ «ЦБ УК г. Сочи»',
         False, 9, '888888'),
        (f'Дата формирования: {today.strftime("%d.%m.%Y")}', False, 9, '888888'),
    ]
    for r_idx, (text, bold, size, color) in enumerate(data, start=2):
        c = ws.cell(row=r_idx, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color or '111111', name='Calibri')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.row_dimensions[r_idx].height = 20 if text else 8


def _write_pfhd_sheet(
    ws,
    data: dict | None,
    all_data: list,
    year: int,
    next_year: int,
    today: date,
    is_summary: bool,
) -> None:
    """Write a ПФХД sheet (either summary or per-institution)."""
    ws.sheet_view.showGridLines = False
    _set_col_widths(ws)

    brd = _thin_border()
    institution = data['institution'] if data else None

    # ── Row 1: Main title ────────────────────────────────────────────────────
    ws.merge_cells('A1:N1')
    c = ws['A1']
    c.value = f'ПЛАН ФИНАНСОВО-ХОЗЯЙСТВЕННОЙ ДЕЯТЕЛЬНОСТИ НА {year} ГОД'
    c.font = Font(bold=True, size=13, color='FFFFFF', name='Calibri')
    c.fill = FILL_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 28

    # ── Row 2: Institution name ──────────────────────────────────────────────
    ws.merge_cells('A2:N2')
    c = ws['A2']
    if is_summary:
        c.value = 'МКУ «Централизованная бухгалтерия УК г. Сочи»  ·  СВОДНЫЙ ПЛАН ПО ВСЕМ ПОДВЕДОМСТВЕННЫМ УЧРЕЖДЕНИЯМ'
    else:
        c.value = institution.name
    c.font = Font(bold=True, size=11, name='Calibri', color='1F3864')
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[2].height = 22

    # ── Row 3: INN / head / legal reference ─────────────────────────────────
    ws.merge_cells('A3:G3')
    if is_summary:
        ws['A3'].value = 'ИНН 2320035700  /  КПП 232001001'
    else:
        ws['A3'].value = f'ИНН: {institution.inn or "—"}'
    ws['A3'].font = FONT_SMALL_GREY
    ws['A3'].alignment = ALIGN_LEFT

    ws.merge_cells('H3:N3')
    if is_summary:
        ws['H3'].value = 'Единица измерения: рублей (коп.)  ·  Приказ Минфина России № 81н от 28.07.2010'
    else:
        ws['H3'].value = f'Руководитель: {institution.head_name or "—"}'
    ws['H3'].font = FONT_SMALL_GREY
    ws['H3'].alignment = Alignment(horizontal='right', vertical='center')
    ws.row_dimensions[3].height = 16

    # ── Row 4: "Единица измерения" (only for per-institution) ────────────────
    ws.merge_cells('A4:N4')
    ws['A4'].value = 'Единица измерения: рублей (коп.)'
    ws['A4'].font = FONT_SMALL_GREY
    ws['A4'].alignment = ALIGN_LEFT
    ws.row_dimensions[4].height = 14

    # ── Rows 5–6: Column headers ─────────────────────────────────────────────
    headers_top = [
        '№', 'КВР', 'КОСГУ', 'Наименование показателя', 'Код\nстр.',
        f'Лимит {year} год, руб.',
        f'Квартальное распределение {year} года, руб.',
        None, None, None,
        'Кассовый расход\nянв.–апр.\nруб.',
        'Исп.\n%',
        f'Лимит {next_year} год, руб.',
        'Изм.\n%',
    ]
    headers_bot = ['1','2','3','4','5','6','I кв.','II кв.','III кв.','IV кв.','11','12','13','14']

    for col_i, val in enumerate(headers_top, start=1):
        c = ws.cell(row=5, column=col_i, value=val)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_COLHEAD
        c.alignment = ALIGN_CENTER
        c.border = brd

    # Merge quarterly distribution header (cols 7-10)
    ws.merge_cells('G5:J5')

    for col_i, val in enumerate(headers_bot, start=1):
        c = ws.cell(row=6, column=col_i, value=val)
        c.font = Font(bold=True, size=9, color='FFFFFF', name='Calibri')
        c.fill = FILL_COLHEAD
        c.alignment = ALIGN_CENTER
        c.border = brd

    ws.row_dimensions[5].height = 42
    ws.row_dimensions[6].height = 14

    # ── Aggregate data ────────────────────────────────────────────────────────
    if is_summary:
        def _sum_field(field: str) -> Decimal:
            return sum(d[field] for d in all_data)

        agg_rev_subsidy = _sum_field('rev_subsidy')
        agg_rev_other   = _sum_field('rev_other')
        agg_rev_paid    = _sum_field('rev_paid')
        agg_total_rev   = _sum_field('total_rev')
        agg_total_fact  = _sum_field('total_fact')

        def _line_amounts(kvr: str, kosgu: str) -> dict:
            result = {k: Decimal('0') for k in ('total','q1','q2','q3','q4')}
            for d in all_data:
                src = d['amounts'].get((kvr, kosgu), {})
                for k in result:
                    result[k] += src.get(k, Decimal('0'))
            return result

        def _fact_for_kosgu(kosgu: str) -> Decimal:
            return sum(d['fact_by_kosgu'].get(kosgu, Decimal('0')) for d in all_data)

        total_limit = _sum_field('total_limit')
        next_limit = total_limit  # same limit for next year in absence of future data
    else:
        agg_rev_subsidy = data['rev_subsidy']
        agg_rev_other   = data['rev_other']
        agg_rev_paid    = data['rev_paid']
        agg_total_rev   = data['total_rev']
        agg_total_fact  = data['total_fact']
        total_limit     = data['total_limit']
        next_limit      = total_limit

        def _line_amounts(kvr: str, kosgu: str) -> dict:
            return data['amounts'].get((kvr, kosgu), {k: Decimal('0') for k in ('total','q1','q2','q3','q4')})

        def _fact_for_kosgu(kosgu: str) -> Decimal:
            return data['fact_by_kosgu'].get(kosgu, Decimal('0'))

    # Revenue totals, other = prochie (5% heuristic for display)
    agg_rev_misc = max(Decimal('0'), agg_total_rev - agg_rev_subsidy - agg_rev_other - agg_rev_paid)

    def _next_limit_for(amount: Decimal, total: Decimal, next_total: Decimal) -> Decimal:
        if total == 0:
            return Decimal('0')
        return (amount / total * next_total).quantize(Decimal('0.01'))

    def _change_pct(cur: Decimal, nxt: Decimal) -> float:
        if cur == 0:
            return 0.0
        return float((nxt - cur) / cur)

    row = 7  # current Excel row

    def style_row(r: int, fill: PatternFill, bold: bool = False, height: int = 16) -> None:
        ws.row_dimensions[r].height = height
        for col in range(1, 15):
            c = ws.cell(row=r, column=col)
            c.fill = fill
            if bold:
                c.font = FONT_BOLD
            c.border = brd

    def write_data(r: int, values: list, bold: bool = False) -> None:
        ws.row_dimensions[r].height = 16
        for col_i, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col_i, value=val)
            c.border = brd
            if bold:
                c.font = FONT_BOLD
            else:
                c.font = FONT_NORMAL
            # Alignment by column
            if col_i in (1, 2, 3, 5):
                c.alignment = ALIGN_CENTER
            elif col_i == 4:
                c.alignment = ALIGN_LEFT
            else:
                c.alignment = ALIGN_RIGHT
            # Number format
            if col_i in range(6, 14) and isinstance(val, (int, float, Decimal)):
                if col_i == 12:
                    c.number_format = NUM_PCT
                elif col_i == 14:
                    c.number_format = NUM_PCT
                else:
                    c.number_format = NUM_RUB

    # ── РАЗДЕЛ 1: ПОСТУПЛЕНИЯ ────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:N{row}')
    c = ws.cell(row=row, column=1, value='РАЗДЕЛ 1. ПОСТУПЛЕНИЯ (ДОХОДЫ)')
    c.font = Font(bold=True, size=10, color='1F3864', name='Calibri')
    c.fill = FILL_SECTION
    c.alignment = ALIGN_LEFT
    c.border = brd
    for col in range(2, 15):
        ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.cell(row=row, column=col).border = brd
    ws.row_dimensions[row].height = 18
    row += 1

    def _rev_row(label: str, code: str, amount: Decimal) -> None:
        nonlocal row
        nxt = _next_limit_for(amount, agg_total_rev, next_limit)
        q = amount / 4
        pct_exec = float(agg_total_fact / amount) if amount else 0.0
        vals = [
            None, None, None, label, code,
            float(amount),
            float(q), float(q), float(q), float(q),
            float(agg_total_fact * (amount / agg_total_rev) if agg_total_rev else 0),
            pct_exec,
            float(nxt),
            _change_pct(amount, nxt),
        ]
        write_data(row, vals)
        row += 1

    _rev_row('Субсидии на выполнение муниципального задания', '010', agg_rev_subsidy)
    _rev_row('Субсидии на иные цели (целевые субсидии)',       '020', agg_rev_other)
    _rev_row('Поступления от приносящей доход деятельности',  '030', agg_rev_paid)
    _rev_row('Прочие поступления',                            '040', agg_rev_misc)

    # ИТОГО поступлений
    nxt_rev = next_limit
    fact_share = agg_total_fact
    vals = [
        'ИТОГО поступлений:', None, None, None, '000',
        float(agg_total_rev),
        float(agg_total_rev / 4), float(agg_total_rev / 4),
        float(agg_total_rev / 4), float(agg_total_rev / 4),
        float(fact_share),
        float(fact_share / agg_total_rev) if agg_total_rev else 0.0,
        float(nxt_rev),
        _change_pct(agg_total_rev, nxt_rev),
    ]
    write_data(row, vals, bold=True)
    ws.merge_cells(f'A{row}:D{row}')
    style_row(row, FILL_TOTAL, bold=True)
    ws.row_dimensions[row].height = 18
    # Re-write the merged cell value
    ws.cell(row=row, column=1).value = 'ИТОГО поступлений:'
    ws.cell(row=row, column=5).value = '000'
    ws.cell(row=row, column=6).value = float(agg_total_rev)
    ws.cell(row=row, column=6).number_format = NUM_RUB
    ws.cell(row=row, column=7).value = float(agg_total_rev / 4)
    ws.cell(row=row, column=7).number_format = NUM_RUB
    ws.cell(row=row, column=8).value = float(agg_total_rev / 4)
    ws.cell(row=row, column=8).number_format = NUM_RUB
    ws.cell(row=row, column=9).value = float(agg_total_rev / 4)
    ws.cell(row=row, column=9).number_format = NUM_RUB
    ws.cell(row=row, column=10).value = float(agg_total_rev / 4)
    ws.cell(row=row, column=10).number_format = NUM_RUB
    ws.cell(row=row, column=11).value = float(fact_share)
    ws.cell(row=row, column=11).number_format = NUM_RUB
    ws.cell(row=row, column=12).value = float(fact_share / agg_total_rev) if agg_total_rev else 0.0
    ws.cell(row=row, column=12).number_format = NUM_PCT
    ws.cell(row=row, column=13).value = float(nxt_rev)
    ws.cell(row=row, column=13).number_format = NUM_RUB
    ws.cell(row=row, column=14).value = _change_pct(agg_total_rev, nxt_rev)
    ws.cell(row=row, column=14).number_format = NUM_PCT
    row += 1

    # ── РАЗДЕЛ 2: ВЫПЛАТЫ ───────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:N{row}')
    c = ws.cell(row=row, column=1, value='РАЗДЕЛ 2. ВЫПЛАТЫ (РАСХОДЫ)')
    c.font = Font(bold=True, size=10, color='1F3864', name='Calibri')
    c.fill = FILL_SECTION
    c.alignment = ALIGN_LEFT
    c.border = brd
    for col in range(2, 15):
        ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.cell(row=row, column=col).border = brd
    ws.row_dimensions[row].height = 18
    row += 1

    line_counter = 0
    kvr_totals: dict = {}   # kvr_code -> {total, q1..q4, fact}
    grand_total = {k: Decimal('0') for k in ('total','q1','q2','q3','q4','fact')}

    # Group lines by KVR group
    kvr_lines: dict = {}
    for (num, kvr_code, kosgu_code, name, row_code) in PFHD_LINES:
        kvr_lines.setdefault(kvr_code, []).append((num, kvr_code, kosgu_code, name, row_code))

    for (kvr_code, group_label) in KVR_GROUPS:
        if kvr_code not in kvr_lines:
            continue

        # KVR group header
        ws.merge_cells(f'A{row}:N{row}')
        c = ws.cell(row=row, column=1, value=group_label)
        c.font = Font(bold=True, size=10, color='2E5E9E', name='Calibri')
        c.fill = FILL_KVR_GRP
        c.alignment = ALIGN_LEFT
        c.border = brd
        for col in range(2, 15):
            ws.cell(row=row, column=col).fill = FILL_KVR_GRP
            ws.cell(row=row, column=col).border = brd
        ws.row_dimensions[row].height = 16
        row += 1

        kvr_sub = {k: Decimal('0') for k in ('total','q1','q2','q3','q4','fact')}

        for (num, kvr_c, kosgu_c, name, row_code) in kvr_lines[kvr_code]:
            line_counter += 1
            am = _line_amounts(kvr_c, kosgu_c)
            fact = _fact_for_kosgu(kosgu_c)
            nxt = _next_limit_for(am['total'], agg_total_rev, next_limit)
            pct_exec = float(fact / am['total']) if am['total'] else 0.0

            vals = [
                line_counter, kvr_c, kosgu_c, name, row_code,
                float(am['total']),
                float(am['q1']), float(am['q2']), float(am['q3']), float(am['q4']),
                float(fact),
                pct_exec,
                float(nxt),
                _change_pct(am['total'], nxt),
            ]
            write_data(row, vals)
            row += 1

            for k in ('total','q1','q2','q3','q4'):
                kvr_sub[k] += am[k]
            kvr_sub['fact'] += fact

        # KVR subtotal
        nxt_sub = _next_limit_for(kvr_sub['total'], agg_total_rev, next_limit)
        pct_sub = float(kvr_sub['fact'] / kvr_sub['total']) if kvr_sub['total'] else 0.0
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1).value = f'Итого по КВР {kvr_code}:'
        ws.cell(row=row, column=6).value  = float(kvr_sub['total'])
        ws.cell(row=row, column=7).value  = float(kvr_sub['q1'])
        ws.cell(row=row, column=8).value  = float(kvr_sub['q2'])
        ws.cell(row=row, column=9).value  = float(kvr_sub['q3'])
        ws.cell(row=row, column=10).value = float(kvr_sub['q4'])
        ws.cell(row=row, column=11).value = float(kvr_sub['fact'])
        ws.cell(row=row, column=12).value = pct_sub
        ws.cell(row=row, column=13).value = float(nxt_sub)
        ws.cell(row=row, column=14).value = _change_pct(kvr_sub['total'], nxt_sub)
        style_row(row, FILL_SUBTOT, bold=True)
        for col in (6,7,8,9,10,11,13):
            ws.cell(row=row, column=col).number_format = NUM_RUB
        ws.cell(row=row, column=12).number_format = NUM_PCT
        ws.cell(row=row, column=14).number_format = NUM_PCT
        row += 1

        for k in ('total','q1','q2','q3','q4','fact'):
            grand_total[k] += kvr_sub[k]

    # ── ИТОГО расходов ───────────────────────────────────────────────────────
    nxt_grand = next_limit
    pct_grand = float(grand_total['fact'] / grand_total['total']) if grand_total['total'] else 0.0
    ws.merge_cells(f'A{row}:D{row}')
    ws.cell(row=row, column=1).value  = 'ИТОГО расходов:'
    ws.cell(row=row, column=5).value  = '200'
    ws.cell(row=row, column=6).value  = float(grand_total['total'])
    ws.cell(row=row, column=7).value  = float(grand_total['q1'])
    ws.cell(row=row, column=8).value  = float(grand_total['q2'])
    ws.cell(row=row, column=9).value  = float(grand_total['q3'])
    ws.cell(row=row, column=10).value = float(grand_total['q4'])
    ws.cell(row=row, column=11).value = float(grand_total['fact'])
    ws.cell(row=row, column=12).value = pct_grand
    ws.cell(row=row, column=13).value = float(nxt_grand)
    ws.cell(row=row, column=14).value = _change_pct(grand_total['total'], nxt_grand)
    style_row(row, FILL_TOTAL, bold=True, height=20)
    for col in (6,7,8,9,10,11,13):
        ws.cell(row=row, column=col).number_format = NUM_RUB
        ws.cell(row=row, column=col).alignment = ALIGN_RIGHT
    ws.cell(row=row, column=12).number_format = NUM_PCT
    ws.cell(row=row, column=14).number_format = NUM_PCT
    row += 1

    # ── РАЗДЕЛ 3: источники ──────────────────────────────────────────────────
    ws.merge_cells(f'A{row}:N{row}')
    ws.cell(row=row, column=1).value = 'РАЗДЕЛ 3. ИСТОЧНИКИ ФИНАНСИРОВАНИЯ ДЕФИЦИТА'
    ws.cell(row=row, column=1).font = Font(bold=True, size=10, color='1F3864', name='Calibri')
    ws.cell(row=row, column=1).fill = FILL_SECTION
    ws.cell(row=row, column=1).alignment = ALIGN_LEFT
    ws.cell(row=row, column=1).border = brd
    for col in range(2, 15):
        ws.cell(row=row, column=col).fill = FILL_SECTION
        ws.cell(row=row, column=col).border = brd
    ws.row_dimensions[row].height = 18
    row += 1

    for label, code in [
        ('Остаток средств на начало года', '300'),
        ('Остаток средств на конец года',  '400'),
        ('Дефицит (+) / Профицит (−)',     '500'),
    ]:
        ws.merge_cells(f'A{row}:D{row}')
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=5).value = code
        ws.cell(row=row, column=6).value = 0
        ws.cell(row=row, column=13).value = 0
        for col in range(1, 15):
            ws.cell(row=row, column=col).border = brd
            ws.cell(row=row, column=col).font = FONT_NORMAL
        ws.row_dimensions[row].height = 16
        row += 1

    # ── Signatures ───────────────────────────────────────────────────────────
    row += 1
    if is_summary:
        sign_lines = [
            f'Директор МКУ «ЦБ УК г. Сочи»: _________________________ / _________________ /',
            f'Главный бухгалтер: _________________________ / _________________ /',
            f'Дата составления: «___» ____________ {year} г.',
        ]
    else:
        sign_lines = [
            f'Руководитель учреждения: _______________________ / {institution.head_name or ""} /',
            f'Главный бухгалтер МКУ «ЦБ УК г. Сочи»: _______________________ /________________/',
            f'Дата: «___» ____________ {year} г.',
        ]
    for line in sign_lines:
        ws.merge_cells(f'A{row}:N{row}')
        c = ws.cell(row=row, column=1, value=line)
        c.font = FONT_NORMAL
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[row].height = 18
        row += 1

    # Freeze panes below header rows
    ws.freeze_panes = 'A7'


def _write_planfact_sheet(ws, all_data: list, year: int) -> None:
    """Write the Plan-vs-Fact summary sheet."""
    ws.sheet_view.showGridLines = False
    brd = _thin_border()

    col_widths = [5, 40, 14, 14, 14, 10]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f'ПЛАН-ФАКТ ИСПОЛНЕНИЯ БЮДЖЕТА — {year} ГОД'
    c.font = Font(bold=True, size=12, color='FFFFFF', name='Calibri')
    c.fill = FILL_TITLE
    c.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 26

    # Header
    headers = ['№', 'Учреждение', 'Лимит (руб.)', 'Согласовано (руб.)', 'Исполнено (руб.)', 'Исп. %']
    for col_i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col_i, value=h)
        c.font = FONT_WHITE_BOLD
        c.fill = FILL_COLHEAD
        c.alignment = ALIGN_CENTER
        c.border = brd
    ws.row_dimensions[2].height = 18

    tot_limit = Decimal('0')
    tot_approved = Decimal('0')
    tot_fact = Decimal('0')

    for idx, data in enumerate(all_data, start=1):
        inst = data['institution']
        limit = data['total_limit']
        # approved = sum of all line totals
        approved = sum(v['total'] for v in data['amounts'].values())
        fact = data['total_fact']
        pct = float(fact / approved) if approved else 0.0

        row = idx + 2
        vals = [idx, inst.short_name, float(limit), float(approved), float(fact), pct]
        for col_i, val in enumerate(vals, start=1):
            c = ws.cell(row=row, column=col_i, value=val)
            c.border = brd
            c.font = FONT_NORMAL
            if col_i == 2:
                c.alignment = ALIGN_LEFT
            elif col_i == 1:
                c.alignment = ALIGN_CENTER
            else:
                c.alignment = ALIGN_RIGHT
            if col_i in (3, 4, 5):
                c.number_format = NUM_RUB
            elif col_i == 6:
                c.number_format = NUM_PCT
        ws.row_dimensions[row].height = 15

        tot_limit    += limit
        tot_approved += approved
        tot_fact     += fact

    # Totals row
    tot_row = len(all_data) + 3
    ws.merge_cells(f'A{tot_row}:B{tot_row}')
    ws.cell(row=tot_row, column=1).value = 'ИТОГО:'
    tot_pct = float(tot_fact / tot_approved) if tot_approved else 0.0
    totals = [None, None, float(tot_limit), float(tot_approved), float(tot_fact), tot_pct]
    for col_i, val in enumerate(totals, start=1):
        c = ws.cell(row=tot_row, column=col_i)
        if val is not None:
            c.value = val
        c.font = FONT_BOLD
        c.fill = FILL_TOTAL
        c.border = brd
        c.alignment = ALIGN_RIGHT
        if col_i in (3, 4, 5):
            c.number_format = NUM_RUB
        elif col_i == 6:
            c.number_format = NUM_PCT
    ws.cell(row=tot_row, column=1).alignment = ALIGN_LEFT
    ws.row_dimensions[tot_row].height = 18
    ws.freeze_panes = 'A3'


# ─── VIEWS ───────────────────────────────────────────────────────────────────

@login_required
def consolidated_report(request):
    """Сводная ведомость: все учреждения × КОСГУ."""
    year = int(request.GET.get('year', 2026))
    institutions = Institution.objects.filter(is_active=True).order_by('short_name')
    kosgu_list = KOSGU.objects.filter(is_active=True).order_by('code')

    rows = []
    for inst in institutions:
        limit_obj = BudgetLimit.objects.filter(institution=inst, year=year).first()
        limit = limit_obj.total_limit if limit_obj else Decimal('0')

        approved_qs = BudgetRequest.objects.filter(
            institution=inst,
            period_year=year,
            status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED],
        )
        approved_total = approved_qs.aggregate(
            total=Sum('approved_amount')
        )['total'] or Decimal('0')

        requested_total = BudgetRequest.objects.filter(
            institution=inst, period_year=year
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        fact_total = FactPayment.objects.filter(
            institution=inst, year=year
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        balance = approved_total - fact_total
        pct = int(fact_total / approved_total * 100) if approved_total else 0

        rows.append({
            'institution': inst,
            'limit': limit,
            'requested': requested_total,
            'approved': approved_total,
            'fact': fact_total,
            'balance': balance,
            'pct': pct,
        })

    totals = {
        'limit': sum(r['limit'] for r in rows),
        'requested': sum(r['requested'] for r in rows),
        'approved': sum(r['approved'] for r in rows),
        'fact': sum(r['fact'] for r in rows),
        'balance': sum(r['balance'] for r in rows),
    }

    return render(request, 'reports/consolidated.html', {
        'rows': rows,
        'totals': totals,
        'year': year,
        'years': range(2024, 2031),
    })


@login_required
def export_excel(request):
    """Экспорт сводной ведомости в Excel (простая форма)."""
    year = int(request.GET.get('year', 2026))
    institutions = Institution.objects.filter(is_active=True).order_by('short_name')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Сводная ведомость {year}'

    header_font = Font(bold=True, size=11)
    title_font  = Font(bold=True, size=13)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = f'МКУ «ЦБ УК г. Сочи» — Сводная ведомость бюджетных заявок на {year} год'
    title_cell.font = title_font
    title_cell.alignment = center

    ws.merge_cells('A2:G2')
    ws['A2'].value = 'Управление культуры Администрации города Сочи'
    ws['A2'].alignment = center

    headers = ['№', 'Учреждение', 'Лимит (руб.)', 'Заявлено (руб.)',
               'Согласовано (руб.)', 'Исполнено (руб.)', 'Остаток (руб.)']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    for col_letter in ('C', 'D', 'E', 'F', 'G'):
        ws.column_dimensions[col_letter].width = 18

    totals = {k: Decimal('0') for k in ('limit', 'requested', 'approved', 'fact', 'balance')}

    for row_idx, inst in enumerate(institutions, start=1):
        limit_obj = BudgetLimit.objects.filter(institution=inst, year=year).first()
        limit = limit_obj.total_limit if limit_obj else Decimal('0')

        requested = BudgetRequest.objects.filter(
            institution=inst, period_year=year
        ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

        approved = BudgetRequest.objects.filter(
            institution=inst, period_year=year,
            status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED]
        ).aggregate(total=Sum('approved_amount'))['total'] or Decimal('0')

        fact = FactPayment.objects.filter(
            institution=inst, year=year
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

        balance = approved - fact
        data_row = [row_idx, inst.short_name, float(limit), float(requested),
                    float(approved), float(fact), float(balance)]

        excel_row = row_idx + 4
        for col_idx, value in enumerate(data_row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.border = border
            if col_idx > 2:
                cell.number_format = '# ##0.00'
                cell.alignment = Alignment(horizontal='right')

        totals['limit']     += limit
        totals['requested'] += requested
        totals['approved']  += approved
        totals['fact']      += fact
        totals['balance']   += balance

    total_row = len(list(institutions)) + 5
    ws.cell(row=total_row, column=1, value='').border = border
    total_label = ws.cell(row=total_row, column=2, value='ИТОГО:')
    total_label.font = header_font
    total_label.border = border

    for col_idx, key in enumerate(('limit', 'requested', 'approved', 'fact', 'balance'), start=3):
        cell = ws.cell(row=total_row, column=col_idx, value=float(totals[key]))
        cell.font = header_font
        cell.number_format = '# ##0.00'
        cell.border = border
        cell.alignment = Alignment(horizontal='right')

    sign_row = total_row + 3
    ws.cell(row=sign_row,     column=1, value='Начальник ЦБ: _______________________')
    ws.cell(row=sign_row + 1, column=1, value='Главный бухгалтер: _______________________')
    ws.cell(row=sign_row + 2, column=1, value=f'Дата: {year} г.')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="budget_report_{year}.xlsx"'
    wb.save(response)
    return response


@login_required
def export_pfhd(request):
    """
    Экспорт ПФХД в многолистовой Excel.
    Структура: Титул → Сводный ПФХД → листы по учреждениям → План-Факт.
    """
    year      = int(request.GET.get('year', 2026))
    next_year = year + 1
    today     = date.today()

    institutions = Institution.objects.filter(is_active=True).order_by('short_name')

    wb = openpyxl.Workbook()

    # Sheet 1: Title page
    ws_title = wb.active
    ws_title.title = 'Титул'
    _write_title_sheet(ws_title, year, today)

    # Collect data for all institutions
    all_data = [_get_institution_pfhd(inst, year) for inst in institutions]

    # Sheet 2: Consolidated ПФХД
    ws_summary = wb.create_sheet('Сводный ПФХД')
    _write_pfhd_sheet(ws_summary, None, all_data, year, next_year, today, is_summary=True)

    # Per-institution sheets
    for data in all_data:
        sheet_name = data['institution'].short_name[:31]
        ws_inst = wb.create_sheet(sheet_name)
        _write_pfhd_sheet(ws_inst, data, [data], year, next_year, today, is_summary=False)

    # Plan-Fact summary sheet
    ws_pf = wb.create_sheet('План-Факт сводная')
    _write_planfact_sheet(ws_pf, all_data, year)

    fname = f'PFHD_MKU_CB_{year}.xlsx'
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    wb.save(response)
    return response
