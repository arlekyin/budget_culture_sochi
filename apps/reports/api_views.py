"""
API views для отчётов.
"""
import io
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from django.http import HttpResponse
from django.db.models import Sum

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated

from apps.institutions.models import Institution, BudgetLimit
from apps.budget_requests.models import BudgetRequest, RequestStatus
from apps.monitoring.models import FactPayment


class ConsolidatedReportAPI(APIView):
    """GET /api/reports/consolidated/?year=2026 — сводная ведомость в JSON."""

    permission_classes = [IsAuthenticated]

    def get(self, request) -> Response:
        year = int(request.query_params.get('year', 2026))
        institutions = Institution.objects.filter(is_active=True).order_by('short_name')

        rows = []
        for inst in institutions:
            limit_obj = BudgetLimit.objects.filter(institution=inst, year=year).first()
            limit = limit_obj.total_limit if limit_obj else Decimal('0')

            requested = BudgetRequest.objects.filter(
                institution=inst, period_year=year
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

            approved = BudgetRequest.objects.filter(
                institution=inst,
                period_year=year,
                status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED],
            ).aggregate(total=Sum('approved_amount'))['total'] or Decimal('0')

            fact = FactPayment.objects.filter(
                institution=inst, year=year
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            balance = approved - fact
            pct = int(fact / approved * 100) if approved else 0

            rows.append({
                'institution_id': inst.id,
                'institution_name': inst.short_name,
                'limit': str(limit),
                'requested': str(requested),
                'approved': str(approved),
                'fact': str(fact),
                'balance': str(balance),
                'pct': pct,
            })

        totals = {
            'limit': str(sum(Decimal(r['limit']) for r in rows)),
            'requested': str(sum(Decimal(r['requested']) for r in rows)),
            'approved': str(sum(Decimal(r['approved']) for r in rows)),
            'fact': str(sum(Decimal(r['fact']) for r in rows)),
            'balance': str(sum(Decimal(r['balance']) for r in rows)),
        }

        return Response({
            'year': year,
            'rows': rows,
            'totals': totals,
        })


class ExportExcelAPI(APIView):
    """GET /api/reports/export/excel/?year=2026 — экспорт в Excel."""

    permission_classes = [IsAuthenticated]

    def get(self, request) -> HttpResponse:
        year = int(request.query_params.get('year', 2026))
        institutions = Institution.objects.filter(is_active=True).order_by('short_name')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Сводная ведомость {year}'

        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=13)
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        header_font_white = Font(bold=True, size=11, color='FFFFFF')
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = (
            f'МКУ «ЦБ УК г. Сочи» — Сводная ведомость бюджетных заявок на {year} год'
        )
        title_cell.font = title_font
        title_cell.alignment = center

        ws.merge_cells('A2:G2')
        subtitle = ws['A2']
        subtitle.value = 'Управление культуры Администрации города Сочи'
        subtitle.alignment = center

        headers = [
            '№', 'Учреждение', 'Лимит (руб.)', 'Заявлено (руб.)',
            'Согласовано (руб.)', 'Исполнено (руб.)', 'Остаток (руб.)'
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 40
        for col_letter in ('C', 'D', 'E', 'F', 'G'):
            ws.column_dimensions[col_letter].width = 18

        totals = {k: Decimal('0') for k in ('limit', 'requested', 'approved', 'fact', 'balance')}

        for row_idx, inst in enumerate(institutions, start=1):
            limit_obj = BudgetLimit.objects.filter(institution=inst, year=year).first()
            limit = limit_obj.total_limit if limit_obj else Decimal('0')

            requested = BudgetRequest.objects.filter(
                institution=inst, period_year=year
            ).aggregate(total=Sum('total_amount'))['total'] or Decimal('0')

            approved = BudgetRequest.objects.filter(
                institution=inst, period_year=year,
                status__in=[RequestStatus.APPROVED, RequestStatus.INCLUDED]
            ).aggregate(total=Sum('approved_amount'))['total'] or Decimal('0')

            fact = FactPayment.objects.filter(
                institution=inst, year=year
            ).aggregate(total=Sum('amount'))['total'] or Decimal('0')

            balance = approved - fact

            data_row = [row_idx, inst.short_name, float(limit), float(requested),
                        float(approved), float(fact), float(balance)]

            excel_row = row_idx + 4
            for col_idx, value in enumerate(data_row, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = border
                if col_idx > 2:
                    cell.number_format = '# ##0.00'
                    cell.alignment = Alignment(horizontal='right')

            totals['limit'] += limit
            totals['requested'] += requested
            totals['approved'] += approved
            totals['fact'] += fact
            totals['balance'] += balance

        inst_count = institutions.count()
        total_row = inst_count + 5
        ws.cell(row=total_row, column=1, value='').border = border
        total_label = ws.cell(row=total_row, column=2, value='ИТОГО:')
        total_label.font = header_font
        total_label.border = border

        for col_idx, key in enumerate(('limit', 'requested', 'approved', 'fact', 'balance'), start=3):
            cell = ws.cell(row=total_row, column=col_idx, value=float(totals[key]))
            cell.font = header_font
            cell.number_format = '# ##0.00'
            cell.border = border
            cell.alignment = Alignment(horizontal='right')

        sign_row = total_row + 3
        ws.cell(row=sign_row, column=1, value='Начальник ЦБ: _______________________')
        ws.cell(row=sign_row + 1, column=1, value='Главный бухгалтер: _______________________')
        ws.cell(row=sign_row + 2, column=1, value=f'Дата: {year} г.')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="budget_report_{year}.xlsx"'
        wb.save(response)
        return response
